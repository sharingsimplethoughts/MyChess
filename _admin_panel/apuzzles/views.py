from django.shortcuts import render
from django.views.generic import TemplateView
from django.http import HttpResponseRedirect
from django.contrib import messages
import openpyxl
import cv2
from puzzles.models import *
from .forms import *


# Create your views here.
class PuzzleCategoriesListView(TemplateView):
    def get(self,request,*args, **kwargs):
        lcs = PuzzleCategory.objects.all()
        for l in lcs:
            count = Puzzle.objects.filter(category=l).count()
            l.count = count
        return render(request,'apuzzles/puzzle-categories.html',{'lcs':lcs})

class PuzzleCategoriesAddView(TemplateView):
    def get(self,request,*args, **kwargs):
        less = Puzzle.objects.filter(is_blocked=False)
        return render(request,'apuzzles/add-puzzle-category.html',{'less':less})
    def post(Self,request,*args,**kwargs):
        less = Puzzle.objects.filter(is_blocked=False)

        context={}
        cr_name = request.POST['cr_name']
        cr_icon = request.FILES.get('cr_icon')
        cr_list = request.POST['cr_list']
        
        context['cr_name']=cr_name
        context['cr_icon']=cr_icon
        form = PuzzleCategoriesAddForm(request.POST or None)
        lcobj=''
        lobjlist=[]
        cr_list = cr_list.split(',')
        for c in cr_list:
            if c:
                lobj = Puzzle.objects.filter(id=c).first()
                lobjlist.append(lobj)
        context['cr_list'] = lobjlist

        if form.is_valid():
            x=0
            if cr_icon:
                if cr_icon.size>5242880:
                    x=1
                    message="Category icon size is not valid"
                if cr_icon.name.split('.')[-1].lower() not in ('jpg','jpeg','png'):
                    x=1
                    message="Category icon extension is not valid"

            if x==1:
                messages.add_message(request, messages.INFO, message)
                return render(request,'apuzzles/add-puzzle-category.html',{'form':form,'less':less,'context':context})


            if cr_icon:
                lcobj = PuzzleCategory(
                    name = cr_name,
                    icon = cr_icon,
                )
                lcobj.save()
                lcobj.iconurl=lcobj.icon.url
                lcobj.save()
            else:
                lcobj = PuzzleCategory(
                    name = cr_name,
                )
                lcobj.save()
            if lobjlist:
                for lobj in lobjlist:
                    if lobj:
                        lobj.category = lcobj
                        lobj.save()
            message = 'Puzzle category created successfully'
            messages.add_message(request, messages.INFO, message)
            return render(request,'apuzzles/add-puzzle-category.html',{'less':less,})

        return render(request,'apuzzles/add-puzzle-category.html',{'form':form,'less':less,'context':context})

class PuzzleCategoriesViewView(TemplateView):
    def get(self,request,*args, **kwargs):
        id = self.kwargs['pk']
        lc = PuzzleCategory.objects.filter(id=id).first()
        lss = Puzzle.objects.filter(category=lc)
        return render(request,'apuzzles/puzzle-category-view.html',{'lc':lc,'lss':lss})

class PuzzleCategoriesEditView(TemplateView):
    def get(self,request,*args, **kwargs):
        id = self.kwargs['pk']
        lc = PuzzleCategory.objects.filter(id=id).first()
        lss = Puzzle.objects.filter(category=lc)

        context={}
        context['cr_id'] = lc.id
        context['cr_name'] = lc.name
        context['cr_icon'] = lc.iconurl
        context['cr_list'] = lss

        less = Puzzle.objects.filter(is_blocked=False)
        return render(request,'apuzzles/puzzle-category-edit.html',{'context':context,'less':less})
    
    def post(self,request, *args, **kwargs):
        id = self.kwargs['pk']
        lc = PuzzleCategory.objects.filter(id=id).first()
        less = Puzzle.objects.filter(is_blocked=False)

        context={}
        
        cr_id = lc.id
        cr_name = request.POST['cr_name']
        cr_icon = request.FILES.get('cr_icon')
        cr_list = request.POST['cr_list']
        
        context['cr_id']=cr_id
        context['cr_name']=cr_name
        context['cr_icon']=cr_icon
        form = PuzzleCategoriesEditForm(request.POST or None, cr_id)
        
        lobjlist=[]
        cr_list = cr_list.split(',')
        for c in cr_list:
            if c:
                lobj = Puzzle.objects.filter(id=c).first()
                lobjlist.append(lobj)
        context['cr_list'] = lobjlist

        if form.is_valid():
            x=0
            if cr_icon:
                if cr_icon.size>5242880:
                    x=1
                    message="Category icon size is not valid"
                if cr_icon.name.split('.')[-1].lower() not in ('jpg','jpeg','png'):
                    x=1
                    message="Category icon extension is not valid"

            if x==1:
                messages.add_message(request, messages.INFO, message)
                return render(request,'apuzzles/puzzle-category-edit.html',{'form':form,'less':less,'context':context})


            if cr_icon:
                lc.name = cr_name
                lc.icon = cr_icon
                lc.save()
                lc.iconurl=lc.icon.url
                lc.save()
            else:
                lc.name = cr_name
                lc.save()
            if lobjlist:
                for lobj in lobjlist:
                    if lobj:
                        lobj.category = lc
                        lobj.save()
            message = 'Puzzle category updated successfully'
            messages.add_message(request, messages.INFO, message)
            return HttpResponseRedirect('/admin_panel/puzzles/categories/view/'+str(cr_id))

        return render(request,'apuzzles/puzzle-category-edit.html',{'form':form,'less':less,'context':context})

class PuzzleCategoriesImportView(TemplateView):
    def post(self,request,*args,**kwargs):
        message=''
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size 

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]
        print(worksheet)
        
        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)

        status = validate_less_cat_excel(excel_data)
        if status!="1":
            messages.add_message(request, messages.INFO, status)
            return HttpResponseRedirect('/admin_panel/puzzles/categories')     
        
        # try:
        counter=0
        for row_data in excel_data:
            counter+=1            
            if counter!=1:
                lesscat = PuzzleCategory(
                    name=row_data[0],
                    iconurl=row_data[1],
                )
                lesscat.save()
                lessids = row_data[2]
                if lessids:
                    if ',' in lessids:
                        lessidslist=lessids.split(',')
                        for v in lessidslist:
                            vobj=Puzzle.objects.filter(id=v).first()
                            vobj.category=lesscat
                            vobj.save()
                    else:
                        if not lessids or lessids == 'None' or lessids=="":
                            pass
                        else:
                            vobj=Puzzle.objects.filter(id=lessids).first()
                            vobj.category=lesscat
                            vobj.save()

                message='Data saved successfully'

        messages.add_message(request, messages.INFO, message)
        return HttpResponseRedirect('/admin_panel/puzzles/categories')

def validate_less_cat_excel(excel_data):
    counter=0
    less_id_list=[]
    if len(excel_data[0])<3:
        return 'Please use proper excel file format'
    for row_data in excel_data:
        counter += 1

        catName=row_data[0]
        catIcon=row_data[1]
        lessIds=row_data[2]
        
        print(counter)
        if counter==1:
            print(catName)
            print(catIcon)
            print(lessIds)
            if (catName!='Category Name' or catIcon!='Category Icon Link' 
                or lessIds!='Puzzle Ids'):
                return 'Please use proper excel file format'
        else:
            if not catName or catName=="" or catName=='None':
                return 'Please provide category name at row '+str(counter)
            if not catIcon or catIcon=="" or catIcon=='None':
                return 'Please provide category icon '+str(counter)

            if ',' in lessIds:
                lessIdsList = lessIds.split(',')
                for v in lessIdsList:
                    vid = Puzzle.objects.filter(id=int(v)).first()
                    less_id_list.append(v)
                    if not vid:
                        return "Puzzle id "+v+" is not correct at row "+str(counter)
            else:
                if not lessIds or lessIds == 'None' or lessIds=="":
                    pass
                else:
                    vid = Puzzle.objects.filter(id=int(lessIds)).first()
                    less_id_list.append(lessIds)
                    if not vid:
                        return "Puzzle id "+lessIds+" is not correct at row "+str(counter)

            if len(catName)>90:
                return "Category name length is exceeding at row "+str(counter)
            if len(catIcon)>8500:
                return "Category Icon length is exceeding at row "+str(counter)
            if 'http' not in catIcon:
                return "Please provide valid category icon link at row "+str(counter)
            
    if counter in (0,1):
        return "Excel file is empty"

    if not len(less_id_list) == len(set(less_id_list)):
        return "Duplicate puzzle ids present"
    return "1"

class PuzzlesListView(TemplateView):
    def get(self,request,*args, **kwargs):
        ls = Puzzle.objects.all()
        return render(request,'apuzzles/puzzle.html',{'ls':ls})

class PuzzlesAddView(TemplateView):
    def get(self,request,*args, **kwargs):
        lcs = PuzzleCategory.objects.filter(is_blocked=False)
        return render(request,'apuzzles/add-puzzle.html',{'lcs':lcs})

    def post(self,request, *args, **kwargs):
        lcs = PuzzleCategory.objects.filter(is_blocked=False)

        les_name = request.POST['les_name']
        les_cat = request.POST['les_cat']
        les_vid = request.FILES.get('les_vid')
        les_desc = request.POST['les_desc']
        les_hint = request.POST['les_hint']
        les_exp = request.POST['les_exp']
        les_learn = request.POST['les_learn']
        context = {
            'les_name':les_name,
            'les_cat':les_cat,
            'les_vid':les_vid,
            'les_desc':les_desc,
            'les_hint':les_hint,
            'les_exp':les_exp,
            'les_learn':les_learn,
        }

        form = PuzzlesAddEditForm(request.POST or None)
        if form.is_valid():
            x=0
            if les_vid:
                # if vidFile.size>10485760:
                #     x=1
                #     message="Video file size is not valid"
                if les_vid.name.split('.')[-1].lower() not in ('mp4','avi','flv','wmv'):
                    x=1
                    message="Video file extension is not valid"
            else:
                x=1
                message="Please upload video file"
            if x==1:
                messages.add_message(request, messages.INFO, message)
                return render(request,'apuzzles/add-puzzle.html',{'lcs':lcs,'context':context})

            obj = PuzzleCategory.objects.filter(id=les_cat).first()
            les = Puzzle(
                name = les_name,
                category = obj,
                video = les_vid,
                description = les_desc,
                hint = les_hint,
                explanation = les_exp,
                learned = les_learn,
            )
            les.save()
            les.videourl = les.video.url
            les.save()
            #Generate Video Preview
            if les.video:
                url=les.video.url
                extension = url.split('.')[-1]
                video_extensions = ['mp4','avi','flv','wmv']
                if extension.lower() in video_extensions:
                    cam = cv2.VideoCapture(les.video.path)
                    if cam.isOpened():
                        ret,frame = cam.read()
                        name = "media_cdn/puzzles/videoprevs/output"+str(les.id)+".jpg"
                        cv2.imwrite(name, frame)
                        les.videopreviewurl="/media/puzzles/videoprevs/output"+str(les.id)+".jpg"
                        les.save()

                        message='Puzzle added successfully'
                        messages.add_message(request, messages.INFO, message)
                        return render(request,'apuzzles/add-puzzle.html',{'lcs':lcs})

        return render(request,'apuzzles/add-puzzle.html',{'form':form,'lcs':lcs, 'context':context})

class PuzzlesViewView(TemplateView):
    def get(self,request,*args, **kwargs):
        id = self.kwargs['pk']
        ls = Puzzle.objects.filter(id=id).first()
        return render(request,'apuzzles/puzzle-view.html',{'ls':ls})

class PuzzlesEditView(TemplateView):
    def get(self,request,*args, **kwargs):
        id = self.kwargs['pk']
        ls = Puzzle.objects.filter(id=id).first()
        lcs = PuzzleCategory.objects.filter(is_blocked=False)

        les_id = ls.id
        les_name = ls.name
        les_cat = ls.category.id
        les_vid = ls.videourl
        les_desc = ls.description
        les_hint = ls.hint
        les_exp = ls.explanation
        les_learn = ls.learned

        context = {
            'les_id':les_id,
            'les_name':les_name,
            'les_cat':les_cat,
            'les_vid':les_vid,
            'les_desc':les_desc,
            'les_hint':les_hint,
            'les_exp':les_exp,
            'les_learn':les_learn,
        }

        return render(request,'apuzzles/puzzle-edit.html',{'context':context,'lcs':lcs})

    def post(self, request, *args, **kwargs):
        id = self.kwargs['pk']
        ls = Puzzle.objects.filter(id=id).first()
        lcs = PuzzleCategory.objects.filter(is_blocked=False)

        les_id = ls.id
        les_name = request.POST['les_name']
        les_cat = request.POST['les_cat']
        les_vid = request.FILES.get('les_vid')
        les_desc = request.POST['les_desc']
        les_hint = request.POST['les_hint']
        les_exp = request.POST['les_exp']
        les_learn = request.POST['les_learn']
        context = {
            'les_id':les_id,
            'les_name':les_name,
            'les_cat':les_cat,
            'les_vid':les_vid,
            'les_desc':les_desc,
            'les_hint':les_hint,
            'les_exp':les_exp,
            'les_learn':les_learn,
        }

        form = PuzzlesAddEditForm(request.POST or None)
        if form.is_valid():
            x=0
            if les_vid:
                # if vidFile.size>10485760:
                #     x=1
                #     message="Video file size is not valid"
                if les_vid.name.split('.')[-1].lower() not in ('mp4','avi','flv','wmv'):
                    x=1
                    message="Video file extension is not valid"
            # else:
            #     x=1
            #     message="Please upload video file"
            if x==1:
                messages.add_message(request, messages.INFO, message)
                return render(request,'apuzzles/puzzle-edit.html',{'context':context,'ls':ls,'lcs':lcs})

            obj = PuzzleCategory.objects.filter(id=les_cat).first()
            
            ls.name=les_name
            ls.category=obj
            if les_vid:
                ls.video=les_vid
            ls.description=les_desc
            ls.hint=les_hint
            ls.explanation=les_exp
            ls.learned=les_learn

            
            ls.save()
            ls.videourl = ls.video.url
            ls.save()
            #Generate Video Preview
            if les_vid:
                url=ls.video.url
                extension = url.split('.')[-1]
                video_extensions = ['mp4','avi','flv','wmv']
                if extension.lower() in video_extensions:
                    cam = cv2.VideoCapture(ls.video.path)
                    if cam.isOpened():
                        ret,frame = cam.read()
                        name = "media_cdn/puzzles/videoprevs/output"+str(ls.id)+".jpg"
                        cv2.imwrite(name, frame)
                        ls.videopreviewurl="/media/puzzles/videoprevs/output"+str(ls.id)+".jpg"
                        ls.save()

            message='Puzzle updated successfully'
            messages.add_message(request, messages.INFO, message)
            return HttpResponseRedirect('/admin_panel/puzzles/list/view/'+str(les_id))

        return render(request,'apuzzles/puzzle-edit.html',{'form':form,'context':context,'ls':ls,'lcs':lcs})

class PuzzlesImportView(TemplateView):
    def post(self,request,*args,**kwargs):
        message=''
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]
        print(worksheet)
        
        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)

        status = validate_excel(excel_data)
        if status!="1":
            messages.add_message(request, messages.INFO, status)
            return HttpResponseRedirect('/admin_panel/puzzles/list')
        
        # try:
        counter=0
        for row_data in excel_data:
            counter+=1            
            if counter!=1:
                cat=PuzzleCategory.objects.filter(id=row_data[1]).first()
                vid = Puzzle(                        
                        name=row_data[0],
                        category=cat,
                        videourl=row_data[2],
                        videopreviewurl=row_data[3],
                        description=row_data[4],
                        hint=row_data[5],
                        explanation=row_data[6],
                        learned=row_data[7],
                    )
                vid.save()                                         
                message='Data saved successfully'
        # except:
            # messages=['Could not save articles..!!',]

        # articles = Article.objects.all().order_by('-created_on')
        messages.add_message(request, messages.INFO, message)
        return HttpResponseRedirect('/admin_panel/puzzles/list')       

def validate_excel(excel_data):
    counter=0    
    if len(excel_data[0])<8:
        return 'Please use proper excel file format'
    for row_data in excel_data:
        counter += 1

        lessName=row_data[0]
        lessCatId=row_data[1]
        vidFileLink=row_data[2]
        vidPreviewLink=row_data[3]
        description=row_data[4]
        hint=row_data[5]
        explanation=row_data[6]
        learned=row_data[7]
        
        print(counter)
        if counter==1:
            if (lessName!='Puzzle Name' or lessCatId!='Category ID' 
            or vidFileLink!='Video Link' or vidPreviewLink!='Video Preview Image Link' 
            or description!='Description' or hint!='Hint' 
            or explanation!='Explanation' or learned!='Learned'):
                return 'Please use proper excel file format'                    
        else:
            if not lessName or lessName=="":
                return 'Please provide puzzle name at row '+str(counter)
            if not lessCatId or lessCatId=="":
                return 'Please provide puzzle category id at row '+str(counter)
            if not vidFileLink or vidFileLink=="":
                return 'Please provide video link at row '+str(counter)
            if not vidPreviewLink or vidPreviewLink=="":
                return 'Please provide video preview link at row '+str(counter)
            if not description or description=="":
                return "Please provide puzzle description at row "+str(counter)
            if not hint or hint=="":
                return "Please provide puzzle hint at row "+str(counter)
            if not explanation or explanation=="":
                return "Please provide puzzle explanation at row "+str(counter)
            if not learned or learned=="":
                return "Please provide puzzle learned text at row "+str(counter)
            if 'http' not in vidFileLink:
                return "Please provide valid video file link at row "+str(counter)
            if 'http' not in vidPreviewLink:
                return "Please provide valid video preview image link at row "+str(counter)

            less=PuzzleCategory.objects.filter(id=lessCatId).first()
            if not less:
                return "Puzzle category id is not valid at row "+str(counter)

            if len(vidFileLink)>8500:
                return "Video file link length is exceeding at row "+str(counter)
            if len(vidPreviewLink)>8500:
                return "Preview image link length is exceeding at row "+str(counter)
            if len(lessName)>300:
                return "Video name length is exceeding at row "+str(counter)
            if len(description)>800:
                return "Puzzle description length is exceeding at row "+str(counter)
            if len(hint)>800:
                return "Puzzle hint length is exceeding at row "+str(counter)
            if len(explanation)>800:
                return "Puzzle explanation length is exceeding at row "+str(counter)
            if len(learned)>800:
                return "Puzzle learned length is exceeding at row "+str(counter)

    if counter in (0,1):
        return "Excel file is empty"
    return "1"


