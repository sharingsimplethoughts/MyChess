from django.shortcuts import render
from django.views.generic import TemplateView, DetailView, ListView, UpdateView, DeleteView
from django.http import HttpResponseRedirect
from django.db.models import Count
from django.contrib import messages
import openpyxl
import cv2
from articles.models import *
from .forms import *
# Create your views here.
class VideosView(TemplateView):
    def get(self,request,*args,**kwargs):
        data=Video.objects.all()        
        return render(request,'avideos/video.html', context={'data':data})

class VideosCategoriesView(TemplateView):
    def get(self,request,*args,**kwargs):
        data=VideoCategory.objects.all().exclude(name='~')
        for d in data:
            count=Video.objects.filter(category=d).count()
            d.count=count
        return render(request,'avideos/video-categories.html',context={'data':data})

class AddVideoView(TemplateView):
    def get(self,request,*args, **kwargs):
        cat=VideoCategory.objects.filter(is_blocked=False).exclude(name='~')
        return render(request,'avideos/add-video.html',context={'cat':cat})
    def post(self,request,*args,**kwargs):
        form = CreateOrEditVideoForm(data=request.POST or None,)
        vidName = request.POST['vidName']
        vidCat = request.POST['status']
        vidAuthName = request.POST['vidAuthName']
        vidAuthCountry = request.POST['vidAuthCountry']
        vidFile = request.FILES.get('vidFile')
        vidAuthImage = request.FILES.get('vidAuthImage')
        vidDesc = request.POST['vidDesc']  

        context={
            'vidName':vidName,
            'vidCat':vidCat,
            'vidAuthName':vidAuthName,
            'vidAuthCountry':vidAuthCountry,
            'vidFile':vidFile,
            'vidDesc':vidDesc,
            'vidAuthImage':vidAuthImage,
        }

        if form.is_valid():            
            x=0
            if vidFile:
                # if vidFile.size>10485760:
                #     x=1
                #     message="Video file size is not valid"
                if vidFile.name.split('.')[-1].lower() not in ('mp4','avi','flv','wmv'):
                    x=1
                    message="Video file extension is not valid"
            else:
                x=1
                message="Please upload video file"
            if vidAuthImage:
                if vidAuthImage.size>5242880:
                    x=1
                    message="Author's image size is not valid"
                if vidAuthImage.name.split('.')[-1].lower() not in ('jpg','jpeg','png'):
                    x=1
                    message="Author's image extension is not valid"

            if x==1:
                cat=VideoCategory.objects.filter(is_blocked=False).exclude(name='~')
                messages.add_message(request, messages.INFO, message)
                return render(request,'avideos/add-video.html',context={'form':form,'cat':cat,'context':context})

            cat=VideoCategory.objects.filter(id=vidCat).first()
            vid=Video(
                name=vidName,
                category=cat,
                vfile=vidFile,
                description=vidDesc,
                authorname=vidAuthName,
                authorpic=vidAuthImage,
                authorcountry=vidAuthCountry,
            )
            vid.save()
            if vid.vfile:
                vid.vfileurl=vid.vfile.url
            if vid.authorpic:
                vid.authorpicurl=vid.authorpic.url
            vid.save()

            ## Generate Preview
            if vid.vfile:
                url=vid.vfile.url            
                extension = url.split('.')[-1]                
                video_extensions = ['mp4','avi','flv','wmv']                
                if extension.lower() in video_extensions:                    
                    cam = cv2.VideoCapture(vid.vfile.path)                    
                    if cam.isOpened():                        
                        ret,frame = cam.read()                        
                        name = "media_cdn/videos/preview/output"+str(vid.id)+".jpg"
                        cv2.imwrite(name, frame)                        
                        vid.vpreviewurl="/media/videos/preview/output"+str(vid.id)+".jpg"                        
                        vid.save()

                        message='Video added successfully'
                        messages.add_message(request, messages.INFO, message)
                        cat=VideoCategory.objects.filter(is_blocked=False).exclude(name='~')
                        return render(request,'avideos/add-video.html',context={'form':form,'cat':cat,})

        cat=VideoCategory.objects.filter(is_blocked=False).exclude(name='~')
        return render(request,'avideos/add-video.html',context={'form':form,'cat':cat,'context':context})

class EditVideoView(TemplateView):
    def get(self,request,*args, **kwargs):
        id=self.kwargs['pk']
        vid=Video.objects.filter(id=id).first()
        cat=VideoCategory.objects.filter(is_blocked=False).exclude(name='~')

        context={
            'vidName':vid.name,
            'vidCat':vid.category.id,
            'vidAuthName':vid.authorname,
            'vidAuthCountry':vid.authorcountry,
            'vidFileUrl':vid.vfileurl,
            'vidDesc':vid.description,
            'vidAuthImageUrl':vid.authorpicurl,
            'vidFile':vid.vfile,
            'vidAuthImage':vid.authorpic,
        }
        return render(request,'avideos/edit-video.html',context={'cat':cat, 'context':context})
    def post(self,request,*args,**kwargs):
        id=self.kwargs['pk']
        vid=Video.objects.filter(id=id).first()

        form = CreateOrEditVideoForm(data=request.POST or None,)
        vidName = request.POST['vidName']
        vidCat = request.POST['status']
        vidAuthName = request.POST['vidAuthName']
        vidAuthCountry = request.POST['vidAuthCountry']
        vidFile = request.FILES.get('vidFile')
        vidAuthImage = request.FILES.get('vidAuthImage')
        vidDesc = request.POST['vidDesc']  

        context={
            'vidName':vidName,
            'vidCat':vidCat,
            'vidAuthName':vidAuthName,
            'vidAuthCountry':vidAuthCountry,
            'vidFile':vidFile,
            'vidDesc':vidDesc,
            'vidAuthImage':vidAuthImage,
        }

        if not vidFile:
            context['vidFileUrl']=vid.vfileurl
        if not vidAuthImage:
            context['vidAuthImageUrl']=vid.authorpicurl

        if form.is_valid():            
            x=0
            if vidFile:
                # if vidFile.size>10485760:
                #     x=1
                #     message="Video file size is not valid"
                if vidFile.name.split('.')[-1].lower() not in ('mp4','avi','flv','wmv'):
                    x=1
                    message="Video file extension is not valid"
            # else:
            #     x=1
            #     message="Please upload video file"
            if vidAuthImage:
                if vidAuthImage.size>5242880:
                    x=1
                    message="Author's image size is not valid"
                if vidAuthImage.name.split('.')[-1].lower() not in ('jpg','jpeg','png'):
                    x=1
                    message="Author's image extension is not valid"

            if x==1:
                cat=VideoCategory.objects.filter(is_blocked=False).exclude(name='~')
                messages.add_message(request, messages.INFO, message)
                return render(request,'avideos/edit-video.html',context={'form':form,'cat':cat,'context':context})

            print(vidCat)
            cat=VideoCategory.objects.filter(id=vidCat).first()
            print(cat) 
            
            vid.name=vidName
            vid.category=cat
            if vidFile:
                vid.vfile=vidFile
            vid.description=vidDesc
            vid.authorname=vidAuthName
            if vidAuthImage:
                vid.authorpic=vidAuthImage
            vid.authorcountry=vidAuthCountry
            
            vid.save()
            if vidFile:
                vid.vfileurl=vid.vfile.url
            if vidAuthImage:
                vid.authorpicurl=vid.authorpic.url
            vid.save()

            ## Generate Preview
            if vidFile:
                url=vid.vfile.url            
                extension = url.split('.')[-1]                
                video_extensions = ['mp4','avi','flv','wmv']                
                if extension.lower() in video_extensions:                    
                    cam = cv2.VideoCapture(vid.vfile.path)                    
                    if cam.isOpened():                        
                        ret,frame = cam.read()                        
                        name = "media_cdn/videos/preview/output"+str(vid.id)+".jpg"
                        cv2.imwrite(name, frame)                        
                        vid.vpreviewurl="/media/videos/preview/output"+str(vid.id)+".jpg"                        
                        vid.save()

            message='Video added successfully'
            messages.add_message(request, messages.INFO, message)
            return HttpResponseRedirect('/admin_panel/videos/view_video/'+str(vid.id))
            # cat=VideoCategory.objects.all()
            # return render(request,'avideos/edit-video.html',context={'form':form,'cat':cat,})

        cat=VideoCategory.objects.filter(is_blocked=False).exclude(name='~')
        return render(request,'avideos/edit-video.html',context={'form':form,'cat':cat,'context':context})

class AddVideoCatView(TemplateView):
    def get(self,request,*args, **kwargs):
        vids=Video.objects.filter(is_blocked=False)
        return render(request,'avideos/add-video-categories.html',context={'vids':vids})

    def post(self,request,*args,**kwargs):
        form = CreateOrEditVideoCategoryForm(data=request.POST or None,)
        catName = request.POST['catName']
        catDesc = request.POST['catDesc']
        finalvideoids = request.POST['finalvideoids']
        # catIcon = request.FILES.get('catIcon')
        vids=Video.objects.filter(is_blocked=False)
        cv1=Video.objects.filter(is_blocked=False).count()
        cv2=''
        ischeckall=0
        selectedvids=''
        if finalvideoids:
            finalvideoidslist = finalvideoids.split(',')
            finalvideoidslist = finalvideoidslist[0:-1]
            print(finalvideoidslist)
            selectedvids=Video.objects.filter(id__in=finalvideoidslist)
            cv2=Video.objects.filter(id__in=finalvideoidslist).count()

        if cv1==cv2:
            ischeckall=1

        context = {
            'catName':catName,
            'catDesc':catDesc,
            'finalvideoids':finalvideoids
        }

        if form.is_valid():
            # x=0
            # if catIcon:
            #     if catIcon.size>5242880:
            #         x=1
            #         message="Category image size is not valid"
            #     if catIcon.name.split('.')[-1].lower() not in ('jpg','jpeg','png'):
            #         x=1
            #         message="Category image extension is not valid"
            # else:
            #     x=1
            #     message="Please upload category icon"
            # if x==1:                
            #     messages.add_message(request, messages.INFO, message)
            #     return render(request,'avideos/add-video-categories.html',context={'form':form,'vids':vids,'context':context,'selectedvids':selectedvids, 'ischeckall':ischeckall})
            cat = VideoCategory(
                name = catName,
                # icon = catIcon,
                description = catDesc,
            )
            cat.save()
            # cat.icon_url = cat.icon.url
            # cat.save()

            if selectedvids:                
                for s in selectedvids:
                    s.category = cat
                    s.is_blocked = cat.is_blocked
                    s.save()
            message='Video category added successfully'
            messages.add_message(request, messages.INFO, message)
            return render(request,'avideos/add-video-categories.html',context={'form':form,'vids':vids})
        
        print('just before end')
        return render(request,'avideos/add-video-categories.html',context={'form':form,'vids':vids,'context':context,'selectedvids':selectedvids, 'ischeckall':ischeckall})

class EditVideoCatView(TemplateView):
    def get(self,request,*args, **kwargs):
        id = self.kwargs['pk']
        cat=VideoCategory.objects.filter(id=id).first()
        selectedvids=Video.objects.filter(category=cat)
        vids=Video.objects.filter(is_blocked=False)
        cv1=Video.objects.filter(is_blocked=False).count()
        cv2=''
        ischeckall=0
        if selectedvids:
            cv2=Video.objects.filter(category=cat).count()
        if cv1==cv2:
            ischeckall=1
        finalvideoids=''
        for s in selectedvids:
            finalvideoids=finalvideoids+str(s.id)+','
        print(finalvideoids)
        context = {
            'catId':cat.id,
            'catName':cat.name,
            # 'catIconUrl':cat.icon_url,
            'catDesc':cat.description,
            'finalvideoids':finalvideoids
        }
        print(context)
        return render(request,'avideos/edit-video-categories.html',context={'vids':vids,'context':context,'selectedvids':selectedvids,'ischeckall':ischeckall})

    def post(self,request,*args,**kwargs):
        id = self.kwargs['pk']
        cat=VideoCategory.objects.filter(id=id).first()
        form = CreateOrEditVideoCategoryForm(data=request.POST or None,)
        catName = request.POST['catName']
        catDesc = request.POST['catDesc']
        finalvideoids = request.POST['finalvideoids']
        catIcon = request.FILES.get('catIcon')
        vids=Video.objects.filter(is_blocked=False)
        cv1=Video.objects.filter(is_blocked=False).count()
        cv2=''
        ischeckall=0
        selectedvids=''
        if finalvideoids:
            finalvideoidslist = finalvideoids.split(',')
            finalvideoidslist = finalvideoidslist[0:-1]
            print(finalvideoidslist)
            selectedvids=Video.objects.filter(id__in=finalvideoidslist)
            cv2=Video.objects.filter(id__in=finalvideoidslist).count()

        if cv1==cv2:
            ischeckall=1

        if catIcon:
            context = {
                'catId':cat.id,
                'catName':catName,
                'catIconUrl':'',
                'catDesc':catDesc,
                'finalvideoids':finalvideoids
            }
        else:
            context = {
                'catId':cat.id,
                'catName':catName,
                'catIconUrl':cat.icon_url,
                'catDesc':catDesc,
                'finalvideoids':finalvideoids
            }
        print(context)
        if form.is_valid():
            # x=0
            # if catIcon:
            #     if catIcon.size>5242880:
            #         x=1
            #         message="Author's image size is not valid"
            #     if catIcon.name.split('.')[-1].lower() not in ('jpg','jpeg','png'):
            #         x=1
            #         message="Author's image extension is not valid"
            
            # if x==1:                
            #     messages.add_message(request, messages.INFO, message)
            #     return render(request,'avideos/edit-video-categories.html',context={'form':form,'vids':vids,'context':context,'selectedvids':selectedvids, 'ischeckall':ischeckall})
            
            cat.name = catName
            if catIcon:
                cat.icon = catIcon
            cat.description = catDesc
            cat.save()
            if catIcon:
                cat.icon_url = cat.icon.url
                cat.save()

            if selectedvids:                
                for s in selectedvids:
                    s.category = cat
                    s.is_blocked = cat.is_blocked
                    s.save()
            allvids=Video.objects.filter(category=cat)
            tempcat=VideoCategory.objects.filter(name='~').first()
            for a in allvids:
                if not selectedvids or selectedvids=="" or a not in selectedvids:
                    a.category=tempcat
                    a.save()
            message='Video category updated successfully'
            messages.add_message(request, messages.INFO, message)
            return HttpResponseRedirect('/admin_panel/videos/view_video_cat/'+str(cat.id))
            # return render(request,'avideos/view-video-category.html',context={'form':form,'vids':vids,'cat':cat})
        
        print('just before end')
        return render(request,'avideos/edit-video-categories.html',context={'form':form,'vids':vids,'context':context,'selectedvids':selectedvids, 'ischeckall':ischeckall})

class ViewVideoView(TemplateView):
    def get(self,request,*args, **kwargs):
        id=self.kwargs['pk']
        vid = Video.objects.filter(id=id).first()
        return render(request,'avideos/view-video.html',context={'vid':vid})

class ViewVideoCatView(TemplateView):
    def get(self,request,*args, **kwargs):
        id=self.kwargs['pk']
        cat = VideoCategory.objects.filter(id=id).first()
        vids = Video.objects.filter(category=cat)
        
        return render(request,'avideos/view-video-category.html',context={'cat':cat, 'vids':vids})

class ImportVideoView(TemplateView):    
    def post(self,request,*args,**kwargs):
        message=''    
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]
        print(worksheet)
        
        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)

        status = validate_excel(excel_data)
        if status!="1":   
            messages.add_message(request, messages.INFO, status)
            return HttpResponseRedirect('/admin_panel/videos/videos')               
        
        # try:
        counter=0
        for row_data in excel_data:
            counter+=1            
            if counter!=1:
                cat=VideoCategory.objects.filter(name=row_data[1]).first()
                vid = Video(                        
                        name=row_data[0],
                        category=cat,
                        vfileurl=row_data[2],
                        vpreviewurl=row_data[3],
                        authorpicurl=row_data[4],
                        authorname=row_data[5],
                        authorcountry=row_data[6],
                        description=row_data[7],
                    )
                vid.save()                                         
                message='Data saved successfully'
        # except:
            # messages=['Could not save articles..!!',]

        # articles = Article.objects.all().order_by('-created_on')
        messages.add_message(request, messages.INFO, message)
        return HttpResponseRedirect('/admin_panel/videos/videos')       

def validate_excel(excel_data):
    counter=0    
    if len(excel_data[0])<8:
        return 'Please use proper excel file format'
    for row_data in excel_data:        
        counter += 1

        vidName=row_data[0]
        vidCatName=row_data[1]
        vidFileLink=row_data[2]
        vidPreviewLink=row_data[3]
        vidAuthImageLink=row_data[4]
        vidAuthName=row_data[5]
        vidAuthCountry=row_data[6]
        vidDesc=row_data[7]
        
        print(counter)
        if counter==1:
            if (vidName!='Video Name' or vidCatName!='Category Name' 
            or vidFileLink!='Video Link' or vidPreviewLink!='Video Preview Image Link' 
            or vidAuthImageLink!='Author Image Link' or vidAuthName!='Author Name' 
            or vidAuthCountry!='Author Country' or vidDesc!='Video Description'):
                return 'Please use proper excel file format'                    
        else:
            if not vidName or vidName=="":
                return 'Please provide video file name at row '+str(counter)
            if not vidCatName or vidCatName=="":
                return 'Please provide video category name at row '+str(counter)
            if not vidFileLink or vidFileLink=="":
                return 'Please provide video link at row '+str(counter)
            if not vidPreviewLink or vidPreviewLink=="":
                return 'Please provide video preview link at row '+str(counter)
            if not vidAuthImageLink or vidAuthImageLink=="":
                return "Please provide video author's image link at row "+str(counter)
            if not vidAuthName or vidAuthName=="":
                return "Please provide author's name at row "+str(counter)
            if not vidAuthCountry or vidAuthCountry=="":
                return "Please provide author's country name at row "+str(counter)
            if not vidDesc or vidDesc=="":
                return "Please provide video description at row "+str(counter)
            if 'http' not in vidFileLink:
                return "Please provide valid video file link at row "+str(counter)
            if 'http' not in vidPreviewLink:
                return "Please provide valid video preview image link at row "+str(counter)
            if 'http' not in vidAuthImageLink:
                return "Please provide valid video author image link at row "+str(counter)

            cat=VideoCategory.objects.filter(name=vidCatName).first()
            if not cat:
                return "Video category name is not valid at row "+str(counter)

            pattern = re.compile("[A-Za-z ]+")
            if pattern.fullmatch(vidAuthName) is None:
                return "Author's name is not valid at row "+str(counter)
            country=None
            country = CountryCode.objects.filter(country__iexact=vidAuthCountry)
            print(country)
            if not country or country=="":
                return "Author's Country is not valid at row "+str(counter)
            if len(vidFileLink)>8500:
                return "Video file link length is exceeding at row "+str(counter)
            if len(vidPreviewLink)>8500:
                return "Preview image link length is exceeding at row "+str(counter)
            if len(vidName)>450:
                return "Video name length is exceeding at row "+str(counter)
            if len(vidDesc)>800:
                return "Article description length is exceeding at row "+str(counter)
            if len(vidAuthName)>15:
                return "Author's name length is exceeding at row "+str(counter)
            if len(vidAuthImageLink)>8500:
                return "Author's image length is exceeding at row "+str(counter)

    if counter in (0,1):
        return "Excel file is empty"
    return "1"


class ImportVideoCategoryView(TemplateView):    
    def post(self,request,*args,**kwargs):
        message=''    
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]
        print(worksheet)
        
        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)

        status = validate_vid_cat_excel(excel_data)
        if status!="1":
            messages.add_message(request, messages.INFO, status)
            return HttpResponseRedirect('/admin_panel/videos/video_categories')     
        
        # try:
        counter=0
        for row_data in excel_data:
            counter+=1            
            if counter!=1:
                vidcat = VideoCategory(
                    name=row_data[0],
                    description=row_data[1],
                    # icon_url=row_data[2],
                )
                vidcat.save()
                vidids = row_data[2]
                if vidids:
                    if ',' in vidids:
                        vididslist=vidids.split(',')
                        for v in vididslist:
                            vobj=Video.objects.filter(id=v).first()
                            vobj.category=vidcat
                            vobj.save()
                    else:
                        if not vidids or vidids == 'None' or vidids=="":
                            pass
                        else:
                            vobj=Video.objects.filter(id=vidids).first()
                            vobj.category=vidcat
                            vobj.save()

                message='Data saved successfully'
        # except:
            # messages=['Could not save articles..!!',]

        # articles = Article.objects.all().order_by('-created_on')
        messages.add_message(request, messages.INFO, message)
        return HttpResponseRedirect('/admin_panel/videos/video_categories')

def validate_vid_cat_excel(excel_data):
    counter=0
    vid_id_list=[]
    if len(excel_data[0])<3:
        return 'Please use proper excel file format'
    for row_data in excel_data:        
        counter += 1

        catName=row_data[0]
        catDesc=row_data[1]
        # catIconLink=row_data[2]
        vidIds=row_data[2]
        
        print(counter)
        if counter==1:
            print(catName)
            print(catDesc)
            print(vidIds)
            if (catName!='Category Name' or catDesc!='Category Description' 
                or vidIds!='Video Ids'):#or catIconLink!='Category Icon Link'
                return 'Please use proper excel file format'                    
        else:
            if not catName or catName=="" or catName=='None':
                print('inside if1')
                return 'Please provide category name at row '+str(counter)
            if not catDesc or catDesc=="" or catDesc=='None':
                return 'Please provide category description '+str(counter)
            # if not catIconLink or catIconLink=="" or catIconLink=='None':
            #     return 'Please provide category icon link '+str(counter)
            # if not vidIds or vidIds=="" or vidIds=='None':
            #     return 'Please provide video ids '+str(counter)
            # if 'http' not in catIconLink:
            #     return "Please provide valid category icon link at row "+str(counter)

            if ',' in vidIds:
                videoIdsList = vidIds.split(',')
                for v in videoIdsList:
                    vid = Video.objects.filter(id=int(v)).first()
                    vid_id_list.append(v)
                    if not vid:
                        return "Video id "+v+" is not correct at row "+str(counter)
            else:
                if not vidIds or vidIds == 'None' or vidIds=="":
                    pass
                else:
                    vid = Video.objects.filter(id=int(vidIds)).first()
                    vid_id_list.append(vidIds)
                    if not vid:
                        return "Video id "+vidIds+" is not correct at row "+str(counter)

            # if len(catIconLink)>8500:
            #     return "Category Icon link length is exceeding at row "+str(counter)
            if len(catName)>90:
                return "Category name length is exceeding at row "+str(counter)
            if len(catDesc)>800:
                return "Category description length is exceeding at row "+str(counter)
            
    if counter in (0,1):
        return "Excel file is empty"

    if not len(vid_id_list) == len(set(vid_id_list)):
        return "Duplicate video ids present"
    return "1"


